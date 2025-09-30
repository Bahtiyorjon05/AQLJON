import asyncio
import tempfile
import logging
import random
import time
from datetime import datetime, timedelta
from telegram import Update
from telegram.ext import ContextTypes
from telegram.constants import ParseMode
from modules.doc_generation.base_generator import BaseDocumentGenerator
from modules.config import Config
from modules.utils import safe_reply, send_typing
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, ScatterChart
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, IconSetRule, Rule
from openpyxl.styles.differential import DifferentialStyle
import io

logger = logging.getLogger(__name__)

class ExcelGenerator(BaseDocumentGenerator):
    """Handles Excel spreadsheet generation with topic-specific content and awesome visuals.
    
    This class generates professional Excel spreadsheets with advanced formatting,
    charts, tables, and conditional formatting. It supports multiple sheets with
    different purposes and provides comprehensive data visualization.
    """
    
    async def generate(self, update: Update, context: ContextTypes.DEFAULT_TYPE, topic: str, content_context: str = ""):
        """Generate a professional Excel spreadsheet based on user request"""
        # Validate topic using centralized validation
        is_valid, cleaned_topic = self._validate_topic(topic)
        if not is_valid:
            await self._track_document_generation(update, "excel")
            if len(cleaned_topic) < 2:
                processing_msg = await self._send_processing_message(update, "<b>❌ Excel hujjat yaratishda xatolik.</b>\n\nMavzu juda qisqa.")
                if processing_msg:
                    await processing_msg.edit_text("❌ Iltimos, Excel jadvali uchun to'liq mavzu kiriting (kamida 2 belgi).", parse_mode=ParseMode.HTML)
            else:
                processing_msg = await self._send_processing_message(update, "<b>❌ Excel hujjat yaratishda xatolik.</b>\n\nIltimos, mavzu kiriting.")
                if processing_msg:
                    await processing_msg.edit_text("❌ Iltimos, Excel jadvali uchun mavzu kiriting.", parse_mode=ParseMode.HTML)
            return
        
        # Track document generation
        await self._track_document_generation(update, "excel")
        
        # Initialize processing_msg to avoid unbound variable error
        processing_msg = None
        
        # Send immediate processing message for better user experience
        try:
            from modules.utils import send_fast_reply
            if update.message:
                send_fast_reply(update.message, "<b>📊 Excel hujjatni tuzyapman. Biroz kutib turing... ⏳</b>")
                # Send typing indicator
                await send_typing(update)
        except:
            # Fallback if fast reply fails
            processing_msg = await self._send_processing_message(update, f"<b>📊 Excel hujjatni tuzyapman. Biroz kutib turing... ⏳</b>")
            # Send typing indicator
            await send_typing(update)
        
        try:
            # Generate filename using centralized method
            filename = await self._generate_filename(cleaned_topic, "Excel")
            
            # Generate content with Gemini using flexible prompt structure - all content in user's language
            # Gemini will automatically detect the language from the user's input
            prompt = f"""
            You are AQLJON, an intelligent assistant who creates exceptional, professionally formatted Excel spreadsheets.
            Create a professional Excel spreadsheet about '{cleaned_topic}' in the SAME LANGUAGE as the user's input.
            
            {"Use the following context from previous documents the user shared to inform your content:" + content_context if content_context else ""}
            
            Analyze the topic and create a well-structured, data-rich Excel file that:
            1. Provides comprehensive coverage based on the specific topic
            2. Uses appropriate data structure for the topic type
            3. Includes realistic sample data with proper formatting
            4. Incorporates advanced Excel features like:
               - Professional styling and formatting with vibrant colors
               - Multiple sheets with specific purposes:
                 * Data sheet: Main data table with comprehensive information
                 * Summary sheet: Key metrics, statistics, and insights
                 * Charts sheet: Visualizations relevant to the topic
                 * Insights sheet: Recommendations and analysis
               - Charts and visualizations relevant to the topic (bar charts, line charts, pie charts, etc.)
               - Tables with filtering and sorting
               - Conditional formatting
               - Summary statistics
               - Proper number formatting (currency, percentages, dates)
               - Auto column width adjustment
            
            IMPORTANT GUIDELINES:
            1. If the topic relates to schedules/routines (like "kun tartibim", "daily routine", "расписание"):
               - Create a time-based schedule table with activities
               - Include time slots, activities, duration, priority
               - Add summary statistics about time allocation
               - Create charts showing time distribution
               
            2. If user asks you to solve academic problems (math, physics, chemistry, coding, biology):
               - NEVER provide direct solutions or answers
               - Guide with concepts, understanding, and hints
               - Focus on educational value and understanding
               - Force them to think and try again while giving understanding clearly
               
            3. For general topics:
               - Provide informative, well-researched data
               - Keep explanations clear and engaging
               - Use appropriate examples and practical insights
               - Include statistics, facts, or data where relevant
               
            5. Never provide illicit, harmful, or inappropriate content
            6. Format the response as CSV data with proper headers for the Data sheet
            7. Include realistic sample data with appropriate data types
            8. Make the content informative, valuable, and professionally structured
            9. Adapt the structure and data based on the topic - be flexible rather than following a rigid template
            10. Create all content in the SAME LANGUAGE as the user's input: '{cleaned_topic}'
            11. Make the content conversational and human-like, avoiding robotic or technical language
            12. Do not include any "Generated by" text or similar robotic phrases
            13. Make the main data sheet as a clean, awesome table with proper styling, colors, and fonts
            14. Include summary statistics and insights in the user's language
            15. Include 2-3 chart visualizations maximum that are relevant to the data
            16. Ensure all categories, headers, and labels are in the user's language
            17. Add emojis and formatting to make the content visually appealing and engaging
            18. For the Summary sheet, provide key metrics like totals, averages, counts, and other relevant statistics
            19. For the Charts sheet, create relevant visualizations based on the data
            20. For the Insights sheet, provide 3-5 actionable insights or recommendations based on the data
            """
            
            # Generate content with timeout to prevent blocking
            try:
                response = await asyncio.wait_for(
                    asyncio.to_thread(lambda: self.model.generate_content(prompt)),
                    timeout=Config.PROCESSING_TIMEOUT * 2  # Double timeout for document generation
                )
                csv_content = response.text if response and response.text else f"Error generating content for {cleaned_topic}."
            except asyncio.TimeoutError:
                logger.warning("Excel content generation timed out, using fallback content")
                csv_content = "Topic,Value\n" + cleaned_topic + ",Data\n"
            
            # Create Excel in temporary file
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
                tmp_path = tmp_file.name
            
            try:
                # Create professional Excel workbook with timeout
                try:
                    wb = await asyncio.wait_for(
                        asyncio.get_event_loop().run_in_executor(
                            None,
                            self._create_excel_workbook,
                            csv_content
                        ),
                        timeout=Config.PROCESSING_TIMEOUT  # Configurable timeout for Excel creation
                    )
                    
                    # Save Excel file
                    wb.save(tmp_path)
                    
                    # Send Excel to user
                    if update.message:
                        with open(tmp_path, 'rb') as excel_file:
                            await update.message.reply_document(
                                document=excel_file,
                                filename=f"{filename}.xlsx",
                                caption=f"📊 <b>'{filename}' mavzusida professional Excel hujjat</b>\n📁 Fayl nomi: {filename}.xlsx",
                                parse_mode=ParseMode.HTML
                            )
                    
                    # Send success message
                    try:
                        from modules.utils import send_fast_reply
                        if update.message:
                            send_fast_reply(update.message, f"✅ <b>'{filename}' nomli Excel hujjatingiz muvaffaqiyatli tuzildi va yuborildi!</b>\n📥 Ajoyib dizayn va did bilan tuzilgan faylingizdan zavqlaning!", parse_mode=ParseMode.HTML)
                    except:
                        # Fallback if fast reply fails
                        if processing_msg is not None:
                            await processing_msg.edit_text(f"✅ <b>'{filename}' nomli Excel hujjatingiz muvaffaqiyatli tuzildi va yuborildi!</b>\n📥 Ajoyib dizayn va did bilan tuzilgan faylingizdan zavqlaning!", parse_mode=ParseMode.HTML)
                
                finally:
                    # Clean up temporary file
                    try:
                        import os
                        if os.path.exists(tmp_path):
                            os.unlink(tmp_path)
                    except Exception as e:
                        logger.warning(f"Failed to cleanup temp file: {e}")
            
            except Exception as e:
                logger.error(f"Excel creation error: {e}")
                raise e
        
        except Exception as e:
            logger.error(f"Excel generation error: {e}")
            # Send error message to user
            try:
                from modules.utils import send_fast_reply
                if update.message:
                    send_fast_reply(update.message, "❌ Excel hujjat yaratishda xatolik. Iltimos, keyinroq qayta urinib ko'ring.", parse_mode=ParseMode.HTML)
            except:
                # Fallback if fast reply fails
                if processing_msg is not None:
                    await processing_msg.edit_text("❌ Excel hujjat yaratishda xatolik. Iltimos, keyinroq qayta urinib ko'ring.", parse_mode=ParseMode.HTML)
    
    def _create_excel_workbook(self, csv_content):
        """Create Excel workbook in a separate thread to avoid blocking asyncio"""
        # Create professional Excel workbook
        wb = Workbook()
        
        # Remove default sheet
        default_sheet = wb.active
        if default_sheet is not None:
            wb.remove(default_sheet)
        
        # Create multiple sheets for professional Excel file
        data_sheet = wb.create_sheet("Ma'lumotlar")
        summary_sheet = wb.create_sheet("Xulosa")
        charts_sheet = wb.create_sheet("Grafiklar")
        insights_sheet = wb.create_sheet("Tushunchalar")
        
        # Ensure sheets were created successfully
        if data_sheet is None:
            raise Exception("Failed to create data sheet")
        
        # Parse CSV content and populate data sheet
        self._parse_and_populate_data(data_sheet, csv_content)
        
        # Add advanced features like tables, conditional formatting, etc.
        if data_sheet is not None:
            self._add_advanced_excel_features(wb, data_sheet)
        
        # Add summary statistics to summary sheet
        if data_sheet is not None and summary_sheet is not None:
            self._add_summary_statistics(wb, summary_sheet, data_sheet)
        
        # Add charts to charts sheet
        if data_sheet is not None and charts_sheet is not None:
            self._add_charts(wb, charts_sheet, data_sheet)
        
        # Add insights to insights sheet
        if data_sheet is not None and insights_sheet is not None:
            self._add_insights(wb, insights_sheet, data_sheet)
        
        # Apply professional styling to all sheets
        self._apply_professional_styling(wb)
        
        return wb
    
    def _parse_and_populate_data(self, data_sheet, csv_content):
        """Parse CSV content and populate the data sheet with vibrant styling"""
        try:
            # Clean up the CSV content
            lines = csv_content.strip().split('\n')
            
            # Remove markdown code block indicators if present
            if lines and (lines[0].startswith("```") or lines[0].startswith("csv")):
                lines = lines[1:]
            if lines and (lines[-1].startswith("```") or lines[-1].startswith("csv")):
                lines = lines[:-1]
            
            # Parse CSV data
            for i, line in enumerate(lines):
                # Handle quoted fields that might contain commas
                cells = []
                in_quotes = False
                current_cell = ""
                for char in line:
                    if char == '"':
                        in_quotes = not in_quotes
                    elif char == ',' and not in_quotes:
                        cells.append(current_cell)
                        current_cell = ""
                    else:
                        current_cell += char
                cells.append(current_cell)  # Add the last cell
                
                # Populate the sheet
                for j, cell in enumerate(cells):
                    # Clean up the cell content and ensure it's a string for headers
                    clean_cell = str(cell).strip().strip('"').strip()
                    # Ensure headers are strings to prevent the openpyxl warning
                    if i == 0:  # First row (headers)
                        clean_cell = str(clean_cell) if clean_cell else f"Ustun_{j+1}"
                    cell_obj = data_sheet.cell(row=i+1, column=j+1, value=clean_cell)
                    
                    # Apply vibrant styling with tons of different colors
                    if i == 0:  # Header row
                        # Use a vibrant color palette for headers
                        header_colors = ["FF6B6B", "4ECDC4", "45B7D1", "96CEB4", "FFEAA7", "DDA0DD", "98D8C8", "F7DC6F", "BB8FCE", "85C1E9"]
                        color_index = j % len(header_colors)
                        header_color = header_colors[color_index]
                        
                        cell_obj.font = Font(bold=True, italic=True, color="FFFFFF", size=14)
                        cell_obj.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
                        cell_obj.alignment = Alignment(horizontal="center", vertical="center")
                        # Add thicker borders for header
                        cell_obj.border = Border(
                            left=Side(style='medium', color='2C3E50'),
                            right=Side(style='medium', color='2C3E50'),
                            top=Side(style='medium', color='2C3E50'),
                            bottom=Side(style='medium', color='2C3E50')
                        )
                    else:  # Data rows
                        cell_obj.font = Font(size=12, bold=(j==1))  # Bold first column (ID)
                        cell_obj.alignment = Alignment(horizontal="left" if j > 1 else "center", vertical="center")
                        # Add alternating row colors with more vibrant colors
                        if (i+1) % 2 == 0:
                            # Use lighter, vibrant colors for even rows
                            row_colors = ["F8F9F9", "E8F4F8", "FFF3E0", "F3E5F5", "E0F7FA", "F1F8E9", "FFF8E1", "E8EAF6"]
                            row_color = row_colors[(i//2) % len(row_colors)]
                            cell_obj.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
                        else:
                            # Use white for odd rows
                            cell_obj.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                        # Add colorful borders
                        border_colors = ["FF6B6B", "4ECDC4", "45B7D1", "96CEB4", "FFEAA7", "DDA0DD"]
                        border_color = border_colors[j % len(border_colors)]
                        cell_obj.border = Border(
                            left=Side(style='thin', color=border_color),
                            right=Side(style='thin', color=border_color),
                            top=Side(style='thin', color=border_color),
                            bottom=Side(style='thin', color=border_color)
                        )
            
            # Auto-adjust column widths
            for column in data_sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                data_sheet.column_dimensions[column_letter].width = min(adjusted_width, 50)
                
        except Exception as e:
            logger.warning(f"Error parsing CSV content: {e}")
            # Generate sample data if parsing fails
            self._generate_sample_data(data_sheet)
    
    def _generate_sample_data(self, data_sheet):
        """Generate sample data when CSV parsing fails with vibrant styling"""
        # Headers - ensure they are all strings with vibrant styling
        headers = ["ID", "Element", "Kategoriya", "Qiymat", "Holat", "Jarayon", "Reyting"]
        # Use a vibrant color palette for headers
        header_colors = ["FF6B6B", "4ECDC4", "45B7D1", "96CEB4", "FFEAA7", "DDA0DD", "98D8C8"]
        
        # Add and style headers with borders
        for col, header in enumerate(headers, 1):
            header_color = header_colors[(col-1) % len(header_colors)]
            cell = data_sheet.cell(row=1, column=col, value=str(header))
            cell.font = Font(bold=True, italic=True, color="FFFFFF", size=14)
            cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # Add thicker borders for header
            cell.border = Border(
                left=Side(style='medium', color='2C3E50'),
                right=Side(style='medium', color='2C3E50'),
                top=Side(style='medium', color='2C3E50'),
                bottom=Side(style='medium', color='2C3E50')
            )
        
        # More realistic sample data based on different topics
        items = ["Loyiha A", "Loyiha B", "Loyiha C", "Loyiha D", "Loyiha E", "Loyiha F", "Loyiha G", "Loyiha H"]
        categories = ["Marketing", "Sotish", "Rivojlanish", "HR", "Moliya", "IT", "Operatsiyalar"]
        statuses = ["Faol", "Kutilmoqda", "Bajarildi", "Bekor qilindi", "Kechildi"]
        
        # Style for data cells with better fonts and colors
        data_font = Font(size=12)
        data_alignment = Alignment(horizontal="left", vertical="center")
        
        # Generate 25 rows of realistic data
        for i in range(2, 27):
            # ID (bold)
            cell = data_sheet.cell(row=i, column=1, value=i-1)
            cell.font = Font(size=12, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Item (Project/Product name)
            cell = data_sheet.cell(row=i, column=2, value=str(random.choice(items)))
            cell.font = data_font
            cell.alignment = data_alignment
            
            # Category
            cell = data_sheet.cell(row=i, column=3, value=str(random.choice(categories)))
            cell.font = data_font
            cell.alignment = data_alignment
            
            # Value (numeric) - more realistic values
            value = round(random.uniform(1000, 50000), 2)
            cell = data_sheet.cell(row=i, column=4, value=value)
            cell.font = Font(size=12, bold=True, color="2C3E50")
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = '#,##0.00'
            
            # Status
            cell = data_sheet.cell(row=i, column=5, value=str(random.choice(statuses)))
            cell.font = data_font
            cell.alignment = data_alignment
            
            # Progress (percentage) - more realistic progression
            progress = round(random.uniform(0, 100), 1)
            cell = data_sheet.cell(row=i, column=6, value=progress)
            cell.font = Font(size=12, bold=True)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = '0.0"%"'
            
            # Rating (1-10 scale for more granularity)
            rating = random.randint(1, 10)
            cell = data_sheet.cell(row=i, column=7, value=rating)
            cell.font = Font(size=12, bold=True, color="E67E22")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Apply alternating row colors and colorful borders to all cells in this row
            # Use lighter, vibrant colors for even rows
            row_colors = ["F8F9F9", "E8F4F8", "FFF3E0", "F3E5F5", "E0F7FA", "F1F8E9", "FFF8E1", "E8EAF6"]
            fill_color = row_colors[((i-2)//2) % len(row_colors)] if (i-1) % 2 == 0 else "FFFFFF"
            
            # Border colors that change for each column
            border_colors = ["FF6B6B", "4ECDC4", "45B7D1", "96CEB4", "FFEAA7", "DDA0DD", "98D8C8"]
            
            for col in range(1, 8):
                cell = data_sheet.cell(row=i, column=col)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                border_color = border_colors[(col-1) % len(border_colors)]
                cell.border = Border(
                    left=Side(style='thin', color=border_color),
                    right=Side(style='thin', color=border_color),
                    top=Side(style='thin', color=border_color),
                    bottom=Side(style='thin', color=border_color)
                )
        
        # Add data validation dropdowns for status column
        from openpyxl.worksheet.datavalidation import DataValidation
        dv = DataValidation(type="list", formula1='"Faol,Kutilmoqda,Bajarildi,Bekor qilindi,Kechildi"', allow_blank=True)
        data_sheet.add_data_validation(dv)
        dv.add(f"E2:E26")
        
        # Add conditional formatting for rating column
        from openpyxl.formatting.rule import ColorScaleRule
        # Color scale for ratings (1-10)
        color_scale_rule = ColorScaleRule(
            start_type="min", start_color="FF0000",
            mid_type="percentile", mid_value=50, mid_color="FFFF00",
            end_type="max", end_color="00FF00"
        )
        data_sheet.conditional_formatting.add("G2:G26", color_scale_rule)
        
        # Auto-adjust column widths
        for column in data_sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            data_sheet.column_dimensions[column_letter].width = min(adjusted_width, 50)
    
    def _add_advanced_excel_features(self, wb, data_sheet):
        """Add advanced Excel features like tables, conditional formatting, etc."""
        try:
            # Add table to data sheet if we have data
            if data_sheet is not None and data_sheet.max_row > 1 and data_sheet.max_column > 1:
                # Ensure all header cells are strings to prevent openpyxl warning
                for col in range(1, data_sheet.max_column + 1):
                    header_cell = data_sheet.cell(row=1, column=col)
                    if header_cell.value is not None:
                        header_cell.value = str(header_cell.value)
                    else:
                        header_cell.value = f"Ustun_{col}"
                
                # Define table range
                max_row = data_sheet.max_row
                max_col = data_sheet.max_column
                column_letters = [get_column_letter(col) for col in range(1, max_col + 1)]
                table_range = f"{column_letters[0]}1:{column_letters[-1]}{max_row}"
                
                # Create table with custom style
                tab = Table(displayName="Ma'lumotlarJadvali", ref=table_range)
                
                # Add a custom table style with vibrant colors
                style = TableStyleInfo(
                    name="TableStyleMedium17",  # Vibrant style
                    showFirstColumn=True,
                    showLastColumn=True,
                    showRowStripes=True,
                    showColumnStripes=True
                )
                tab.tableStyleInfo = style
                data_sheet.add_table(tab)
                
                # Add auto filter
                data_sheet.auto_filter.ref = table_range
                
                # Add conditional formatting for numeric columns
                # Color scale for Value column (column 4)
                if max_col >= 4:
                    value_column = get_column_letter(4)
                    color_scale_rule = ColorScaleRule(
                        start_type="min", start_color="F8696B",  # Red
                        mid_type="percentile", mid_value=50, mid_color="FFEB84",  # Yellow
                        end_type="max", end_color="63BE7B"  # Green
                    )
                    data_sheet.conditional_formatting.add(f"{value_column}2:{value_column}{max_row}", color_scale_rule)
                
                # Data bars for Progress column (column 6)
                if max_col >= 6:
                    progress_column = get_column_letter(6)
                    data_bar_rule = DataBarRule(
                        start_type="min", end_type="max",
                        color="63BE7B",  # Green
                        showValue=True, minLength=None, maxLength=None
                    )
                    data_sheet.conditional_formatting.add(f"{progress_column}2:{progress_column}{max_row}", data_bar_rule)
                
                # Icon set for Rating column (column 7)
                if max_col >= 7:
                    rating_column = get_column_letter(7)
                    icon_set_rule = IconSetRule(
                        '3Arrows', 'percent', [0, 33, 67],
                        showValue=True, percent=True, reverse=False
                    )
                    data_sheet.conditional_formatting.add(f"{rating_column}2:{rating_column}{max_row}", icon_set_rule)
                
                # Add conditional formatting for Status column (column 5)
                if max_col >= 5:
                    status_column = get_column_letter(5)
                    # Highlight "Active" status in green
                    active_rule = Rule(
                        type="containsText", 
                        formula=[f'NOT(ISERROR(SEARCH("Faol",{status_column}1)))'],
                        stopIfTrue=True
                    )
                    active_dxf = DifferentialStyle(
                        fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                        font=Font(color="006100")
                    )
                    active_rule.dxf = active_dxf
                    data_sheet.conditional_formatting.add(f"{status_column}2:{status_column}{max_row}", active_rule)
                    
                    # Highlight "Completed" status in blue
                    completed_rule = Rule(
                        type="containsText", 
                        formula=[f'NOT(ISERROR(SEARCH("Bajarildi",{status_column}1)))'],
                        stopIfTrue=True
                    )
                    completed_dxf = DifferentialStyle(
                        fill=PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),
                        font=Font(color="00008B")
                    )
                    completed_rule.dxf = completed_dxf
                    data_sheet.conditional_formatting.add(f"{status_column}2:{status_column}{max_row}", completed_rule)
                    
                    # Highlight "Pending" status in yellow
                    pending_rule = Rule(
                        type="containsText", 
                        formula=[f'NOT(ISERROR(SEARCH("Kutilmoqda",{status_column}1)))'],
                        stopIfTrue=True
                    )
                    pending_dxf = DifferentialStyle(
                        fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
                        font=Font(color="9C5700")
                    )
                    pending_rule.dxf = pending_dxf
                    data_sheet.conditional_formatting.add(f"{status_column}2:{status_column}{max_row}", pending_rule)
                    
                    # Highlight "Cancelled" status in red
                    cancelled_rule = Rule(
                        type="containsText", 
                        formula=[f'NOT(ISERROR(SEARCH("Bekor qilindi",{status_column}1)))'],
                        stopIfTrue=True
                    )
                    cancelled_dxf = DifferentialStyle(
                        fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                        font=Font(color="9C0006")
                    )
                    cancelled_rule.dxf = pending_dxf
                    data_sheet.conditional_formatting.add(f"{status_column}2:{status_column}{max_row}", cancelled_rule)
                
                # Add sparklines for numeric data
                # This would be a great addition but requires more complex implementation
                # For now, we'll focus on what we have
                
        except Exception as e:
            logger.warning(f"Error adding advanced Excel features: {e}")

    def _add_summary_statistics(self, wb, summary_sheet, data_sheet):
        """Add summary statistics to the summary sheet with enhanced styling"""
        try:
            # Style the summary sheet header with gradient fill
            summary_sheet['A1'] = "📊 Xulosa va Statistikalar"
            summary_sheet['A1'].font = Font(size=24, bold=True, color="FFFFFF", italic=True)
            summary_sheet['A1'].fill = PatternFill(start_color="4ECDC4", end_color="96CEB4", fill_type="solid")
            summary_sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")
            summary_sheet.merge_cells('A1:E1')
            
            # Add decorative separator with gradient
            for col in range(1, 6):
                cell = summary_sheet.cell(row=2, column=col)
                # Create gradient effect
                gradient_colors = ["2C3E50", "34495E", "2C3E50"]
                color_index = (col - 1) % len(gradient_colors)
                cell.fill = PatternFill(start_color=gradient_colors[color_index], end_color=gradient_colors[color_index], fill_type="solid")
            
            # Add summary statistics
            if data_sheet is not None and data_sheet.max_row > 1 and data_sheet.max_column > 1:
                # Total rows
                summary_sheet['A3'] = "📈 Jami qatorlar:"
                summary_sheet['B3'] = data_sheet.max_row - 1
                summary_sheet['A3'].font = Font(bold=True, size=14, color="2C3E50")
                summary_sheet['B3'].font = Font(bold=True, color="3498DB", size=14)
                summary_sheet['B3'].number_format = '#,##0'
                
                # Total columns
                summary_sheet['A4'] = "📋 Jami ustunlar:"
                summary_sheet['B4'] = data_sheet.max_column
                summary_sheet['A4'].font = Font(bold=True, size=14, color="2C3E50")
                summary_sheet['B4'].font = Font(bold=True, color="3498DB", size=14)
                
                # Add more detailed statistics
                summary_sheet['A6'] = "📊 Batafsil statistikalar:"
                summary_sheet['A6'].font = Font(bold=True, size=18, underline="single", color="45B7D1")
                
                # Numeric column statistics
                numeric_columns = []
                text_columns = []
                
                for col in range(1, min(data_sheet.max_column + 1, 10)):  # Check first 10 columns
                    try:
                        # Check if column contains numeric data
                        is_numeric = True
                        is_text = True
                        values = []
                        
                        for row in range(2, min(data_sheet.max_row + 1, 20)):  # Check first 20 rows
                            cell_value = data_sheet.cell(row=row, column=col).value
                            if cell_value is not None:
                                try:
                                    float(cell_value)
                                    values.append(float(cell_value))
                                    is_text = False
                                except (ValueError, TypeError):
                                    is_numeric = False
                                    is_text = True
                        
                        header = data_sheet.cell(row=1, column=col).value or f"Ustun {get_column_letter(col)}"
                        
                        if is_numeric and values:
                            numeric_columns.append((col, header, values))
                        elif is_text:
                            text_columns.append((col, header))
                    except:
                        pass
                
                # Add statistics for numeric columns
                row_offset = 7
                for col_index, header, values in numeric_columns[:3]:  # Show max 3 numeric columns
                    if len(values) > 0:
                        col_letter = get_column_letter(col_index)
                        
                        # Header for this column with styling
                        summary_sheet[f'A{row_offset}'] = f"📊 {header}:"
                        summary_sheet[f'A{row_offset}'].font = Font(bold=True, italic=True, size=15, color="2C3E50")
                        # Add background color
                        summary_sheet[f'A{row_offset}'].fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
                        row_offset += 1
                        
                        # Statistics with emojis and better formatting
                        summary_sheet[f'A{row_offset}'] = "  📏 O'rtacha:"
                        summary_sheet[f'B{row_offset}'] = sum(values) / len(values)
                        summary_sheet[f'B{row_offset}'].number_format = '#,##0.00'
                        summary_sheet[f'B{row_offset}'].font = Font(color="45B7D1", bold=True, size=12)
                        row_offset += 1
                        
                        summary_sheet[f'A{row_offset}'] = "  🔽 Minimal:"
                        summary_sheet[f'B{row_offset}'] = min(values)
                        summary_sheet[f'B{row_offset}'].number_format = '#,##0.00'
                        summary_sheet[f'B{row_offset}'].font = Font(color="E74C3C", bold=True, size=12)
                        row_offset += 1
                        
                        summary_sheet[f'A{row_offset}'] = "  🔼 Maksimal:"
                        summary_sheet[f'B{row_offset}'] = max(values)
                        summary_sheet[f'B{row_offset}'].number_format = '#,##0.00'
                        summary_sheet[f'B{row_offset}'].font = Font(color="27AE60", bold=True, size=12)
                        row_offset += 1
                        
                        summary_sheet[f'A{row_offset}'] = "  💰 Jami:"
                        summary_sheet[f'B{row_offset}'] = sum(values)
                        summary_sheet[f'B{row_offset}'].number_format = '#,##0.00'
                        summary_sheet[f'B{row_offset}'].font = Font(color="F39C12", bold=True, size=12)
                        row_offset += 2  # Extra space
                
                # Add category analysis for text columns
                if text_columns:
                    summary_sheet[f'A{row_offset}'] = "🏷️ Kategoriya tahlili:"
                    summary_sheet[f'A{row_offset}'].font = Font(bold=True, size=18, underline="single", color="96CEB4")
                    # Add background color
                    summary_sheet[f'A{row_offset}'].fill = PatternFill(start_color="F8F9F9", end_color="F8F9F9", fill_type="solid")
                    row_offset += 1
                    
                    for col_index, header in text_columns[:2]:  # Show max 2 text columns
                        # Count unique values
                        values = []
                        for row in range(2, min(data_sheet.max_row + 1, 20)):
                            cell_value = data_sheet.cell(row=row, column=col_index).value
                            if cell_value is not None:
                                values.append(str(cell_value))
                        
                        if values:
                            unique_count = len(set(values))
                            most_common = max(set(values), key=values.count) if values else "N/A"
                            
                            summary_sheet[f'A{row_offset}'] = f"📌 {header}:"
                            summary_sheet[f'A{row_offset}'].font = Font(bold=True, italic=True, size=15, color="2C3E50")
                            # Add background color
                            summary_sheet[f'A{row_offset}'].fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
                            row_offset += 1
                            
                            summary_sheet[f'A{row_offset}'] = "  🎯 Noyob qiymatlar:"
                            summary_sheet[f'B{row_offset}'] = unique_count
                            summary_sheet[f'B{row_offset}'].font = Font(color="9B59B6", bold=True, size=12)
                            row_offset += 1
                            
                            summary_sheet[f'A{row_offset}'] = "  🔝 Eng ko'p uchraydigan:"
                            summary_sheet[f'B{row_offset}'] = most_common
                            summary_sheet[f'B{row_offset}'].font = Font(color="3498DB", bold=True, size=12)
                            row_offset += 2  # Extra space
                
                # Add data quality indicators
                summary_sheet[f'A{row_offset}'] = "🔍 Ma'lumotlar sifati:"
                summary_sheet[f'A{row_offset}'].font = Font(bold=True, size=18, underline="single", color="FFEAA7")
                # Add background color
                summary_sheet[f'A{row_offset}'].fill = PatternFill(start_color="F8F9F9", end_color="F8F9F9", fill_type="solid")
                row_offset += 1
                
                # Calculate data completeness
                total_cells = (data_sheet.max_row - 1) * data_sheet.max_column
                filled_cells = 0
                for row in range(2, data_sheet.max_row + 1):
                    for col in range(1, data_sheet.max_column + 1):
                        if data_sheet.cell(row=row, column=col).value is not None:
                            filled_cells += 1
                
                completeness = (filled_cells / total_cells) * 100 if total_cells > 0 else 0
                
                summary_sheet[f'A{row_offset}'] = "  📦 To'ldirilgan:"
                summary_sheet[f'B{row_offset}'] = f"{completeness:.1f}%"
                # Color code based on completeness
                if completeness > 80:
                    color = "27AE60"  # Green
                elif completeness > 50:
                    color = "F39C12"  # Orange
                else:
                    color = "E74C3C"  # Red
                summary_sheet[f'B{row_offset}'].font = Font(bold=True, color=color, size=14)
                # Add background color
                summary_sheet[f'A{row_offset}'].fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
                row_offset += 1
                
                # Add timestamp
                summary_sheet[f'A{row_offset}'] = "⏱️ Yaratilgan vaqt:"
                summary_sheet[f'B{row_offset}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                summary_sheet[f'B{row_offset}'].font = Font(italic=True, color="7F8C8D", size=12)
                # Add background color
                summary_sheet[f'A{row_offset}'].fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
            
            # Style the summary sheet with borders and better formatting
            for row in summary_sheet.iter_rows(min_row=1, max_row=summary_sheet.max_row, min_col=1, max_col=summary_sheet.max_column):
                for cell in row:
                    if cell.row > 1 and cell.value is not None:
                        # Apply gradient border effect
                        border_colors = ["4ECDC4", "96CEB4", "FFEAA7", "DDA0DD"]
                        color_index = (cell.row + cell.column) % len(border_colors)
                        border_color = border_colors[color_index]
                        
                        cell.border = Border(
                            left=Side(style='thin', color=border_color),
                            right=Side(style='thin', color=border_color),
                            top=Side(style='thin', color=border_color),
                            bottom=Side(style='thin', color=border_color)
                        )
            
            # Auto-adjust column widths
            for column in summary_sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                summary_sheet.column_dimensions[column_letter].width = min(adjusted_width, 50)
                
        except Exception as e:
            logger.warning(f"Error adding summary statistics: {e}")

    def _add_charts(self, wb, charts_sheet, data_sheet):
        """Add charts to the charts sheet with enhanced styling"""
        try:
            # Style the charts sheet header with gradient fill
            charts_sheet['A1'] = "📊 Grafiklar va Vizualizatsiyalar"
            charts_sheet['A1'].font = Font(size=20, bold=True, color="FFFFFF")
            charts_sheet['A1'].fill = PatternFill(start_color="96CEB4", end_color="45B7D1", fill_type="solid")
            charts_sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")
            charts_sheet.merge_cells('A1:J1')
            
            # Add decorative separator
            for col in range(1, 11):
                cell = charts_sheet.cell(row=2, column=col)
                cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            
            # Add charts if we have data
            if data_sheet is not None and data_sheet.max_row > 2 and data_sheet.max_column > 2:
                # Create a bar chart with enhanced styling
                bar_chart = None
                try:
                    bar_chart = BarChart()
                    bar_chart.title = "📊 Ma'lumotlar taqsimoti"
                    bar_chart.x_axis.title = "Kategoriyalar"
                    bar_chart.y_axis.title = "Qiymatlar"
                    bar_chart.grouping = "clustered"
                    bar_chart.overlap = 10
                    bar_chart.gapWidth = 50
                except Exception as e:
                    logger.warning(f"Error creating bar chart: {e}")
                
                # Set data for bar chart if creation was successful
                if bar_chart is not None:
                    try:
                        categories = Reference(data_sheet, min_col=2, min_row=2, max_row=min(data_sheet.max_row, 15))
                        values = Reference(data_sheet, min_col=4, min_row=1, max_row=min(data_sheet.max_row, 15))
                        bar_chart.add_data(values, titles_from_data=True)
                        bar_chart.set_categories(categories)
                    except Exception as e:
                        logger.warning(f"Error setting data for bar chart: {e}")
                        bar_chart = None
                
                # Add data labels if chart was created successfully
                if bar_chart is not None:
                    bar_chart.dataLabels = DataLabelList()
                    bar_chart.dataLabels.showVal = True
                    bar_chart.dataLabels.showCatName = False
                
                # Style the bar chart
                if bar_chart is not None:
                    bar_chart.legend.position = 'b'
                    bar_chart.style = 10
                    # Add more visual enhancements
                    bar_chart.overlap = 20
                    bar_chart.gapWidth = 30
                
                # Position and size the chart
                if bar_chart is not None:
                    bar_chart.width = 15
                    bar_chart.height = 8
                
                # Add the bar chart to the sheet
                if bar_chart is not None:
                    charts_sheet.add_chart(bar_chart, "A3")
                
                # Create a pie chart if we have categorical data
                pie_chart = None
                try:
                    pie_chart = PieChart()
                    pie_chart.title = "🥧 Kategoriya taqsimoti"
                except Exception as e:
                    logger.warning(f"Error creating pie chart: {e}")
                
                # Set data for pie chart if creation was successful
                if pie_chart is not None:
                    try:
                        categories = Reference(data_sheet, min_col=3, min_row=2, max_row=min(data_sheet.max_row, 12))
                        values = Reference(data_sheet, min_col=4, min_row=2, max_row=min(data_sheet.max_row, 12))
                        pie_chart.add_data(values)
                        pie_chart.set_categories(categories)
                    except Exception as e:
                        logger.warning(f"Error setting data for pie chart: {e}")
                        pie_chart = None
                
                # Add data labels if chart was created successfully
                if pie_chart is not None:
                    pie_chart.dataLabels = DataLabelList()
                    pie_chart.dataLabels.showVal = True
                    pie_chart.dataLabels.showCatName = True
                    pie_chart.dataLabels.showPercent = True
                
                # Style the pie chart
                if pie_chart is not None:
                    if pie_chart.legend:
                        pie_chart.legend.position = 'b'
                    pie_chart.style = 12
                    # Add more visual enhancements
                    pie_chart.varyColors = True
                
                # Position and size the chart
                if pie_chart is not None:
                    pie_chart.width = 12
                    pie_chart.height = 8
                
                # Add the pie chart to the sheet
                if pie_chart is not None:
                    charts_sheet.add_chart(pie_chart, "A15")
                
                # Create a line chart if we have time-series data
                line_chart = None
                try:
                    line_chart = LineChart()
                    line_chart.title = "📈 Trendlar"
                    line_chart.x_axis.title = "Vaqt"
                    line_chart.y_axis.title = "Qiymatlar"
                    line_chart.grouping = "standard"
                except Exception as e:
                    logger.warning(f"Error creating line chart: {e}")
                
                # Set data for line chart if creation was successful
                if line_chart is not None:
                    try:
                        categories = Reference(data_sheet, min_col=2, min_row=2, max_row=min(data_sheet.max_row, 18))
                        values = Reference(data_sheet, min_col=6, min_row=1, max_row=min(data_sheet.max_row, 18))
                        line_chart.add_data(values, titles_from_data=True)
                        line_chart.set_categories(categories)
                    except Exception as e:
                        logger.warning(f"Error setting data for line chart: {e}")
                        line_chart = None
                
                # Add data labels if chart was created successfully
                if line_chart is not None:
                    line_chart.dataLabels = DataLabelList()
                    line_chart.dataLabels.showVal = True
                    line_chart.dataLabels.showCatName = False
                
                # Style the line chart
                if line_chart is not None:
                    if line_chart.legend:
                        line_chart.legend.position = 'b'
                    line_chart.style = 14
                    # Add more visual enhancements
                    line_chart.smooth = True
                
                # Position and size the chart
                if line_chart is not None:
                    line_chart.width = 15
                    line_chart.height = 8
                
                # Add the line chart to the sheet
                if line_chart is not None:
                    charts_sheet.add_chart(line_chart, "F3")
                
                # Create a scatter chart for correlation analysis
                scatter_chart = None
                try:
                    scatter_chart = ScatterChart()
                    scatter_chart.title = "📉 Korrelyatsiya tahlili"
                    scatter_chart.x_axis.title = "Qiymatlar"
                    scatter_chart.y_axis.title = "Jarayon"
                except Exception as e:
                    logger.warning(f"Error creating scatter chart: {e}")
                
                # Set data for scatter chart if creation was successful
                if scatter_chart is not None:
                    try:
                        x_values = Reference(data_sheet, min_col=4, min_row=2, max_row=min(data_sheet.max_row, 15))
                        y_values = Reference(data_sheet, min_col=6, min_row=2, max_row=min(data_sheet.max_row, 15))
                        scatter_chart.add_data(y_values, titles_from_data=False)
                        scatter_chart.series[0].xvalues = x_values
                    except Exception as e:
                        logger.warning(f"Error setting data for scatter chart: {e}")
                        scatter_chart = None
                
                # Style the scatter chart
                if scatter_chart is not None:
                    if scatter_chart.legend:
                        scatter_chart.legend.position = 'b'
                    scatter_chart.style = 15
                
                # Position and size the chart
                if scatter_chart is not None:
                    scatter_chart.width = 12
                    scatter_chart.height = 8
                
                # Add the scatter chart to the sheet
                if scatter_chart is not None:
                    charts_sheet.add_chart(scatter_chart, "F15")
            
            # Auto-adjust column widths
            for column in charts_sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                charts_sheet.column_dimensions[column_letter].width = min(adjusted_width, 50)
                
        except Exception as e:
            logger.warning(f"Error adding charts: {e}")

    def _add_insights(self, wb, insights_sheet, data_sheet):
        """Add insights to the insights sheet with enhanced styling"""
        try:
            # Style the insights sheet header with gradient fill
            insights_sheet['A1'] = "💡 Tushunchalar va Tavsiyalar"
            insights_sheet['A1'].font = Font(size=20, bold=True, color="FFFFFF")
            insights_sheet['A1'].fill = PatternFill(start_color="FFEAA7", end_color="FF6B6B", fill_type="solid")
            insights_sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")
            insights_sheet.merge_cells('A1:E1')
            
            # Add decorative separator
            for col in range(1, 6):
                cell = insights_sheet.cell(row=2, column=col)
                cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            
            # Add comprehensive insights based on data analysis
            insights_sheet['A3'] = "🎯 1. Asosiy tushunchalar:"
            insights_sheet['A3'].font = Font(bold=True, size=14, color="2C3E50")
            
            insights_sheet['A5'] = "📊 Ma'lumotlar tahlili asosida, eng yuqori qiymatlar aniq kategoriyalarda jamlangan."
            insights_sheet['A6'] = "📈 Jarayon ko'rsatkichlari umumiy tendentsiyaga mos keladi."
            insights_sheet['A7'] = "⭐ Reytinglar tarqoq, bu turli xil natijalarni ko'rsatadi."
            insights_sheet['A8'] = "📋 Ma'lumotlar to'plami muvozanatli va tushunarli tuzilgan."
            
            insights_sheet['A10'] = "✅ 2. Tavsiyalar:"
            insights_sheet['A10'].font = Font(bold=True, size=14, color="27AE60")
            
            insights_sheet['A12'] = "🔄 Ma'lumotlarni muntazam ravishda yangilab turing."
            insights_sheet['A13'] = "🔍 Kategoriyalarni yanada aniqlashtirish tavsiya etiladi."
            insights_sheet['A14'] = "📈 Qo'shimcha statistik tahlillar o'tkazish foydali bo'lishi mumkin."
            insights_sheet['A15'] = "📊 Grafiklarni muntazam yangilash orqali tendentsiyalarni kuzatib boring."
            
            insights_sheet['A17'] = "🚀 3. Keyingi qadamlar:"
            insights_sheet['A17'].font = Font(bold=True, size=14, color="3498DB")
            
            insights_sheet['A19'] = "📂 Yangi ma'lumotlar bazasini yaratish."
            insights_sheet['A20'] = "📈 Mavjud ma'lumotlarni kengaytirish."
            insights_sheet['A21'] = "🎨 Grafiklarni yangilash va takomillashtirish."
            insights_sheet['A22'] = "🔍 Ma'lumotlar sifatini yaxshilash uchun tekshiruvlar o'tkazish."
            
            # Add data-driven insights section
            insights_sheet['A24'] = "🔬 4. Ma'lumotlar tahlili:"
            insights_sheet['A24'].font = Font(bold=True, size=14, color="9B59B6")
            
            # Analyze the data to provide specific insights
            if data_sheet is not None and data_sheet.max_row > 2:
                # Sample data analysis
                insights_sheet['A26'] = f"📋 Ma'lumotlar to'plami o'rtacha hajmi: {data_sheet.max_row - 1} qator"
                insights_sheet['A27'] = f"📊 Ma'lumotlar to'plamida {data_sheet.max_column} ta ustun mavjud"
                
                # Check for numeric data
                numeric_cols = 0
                for col in range(1, data_sheet.max_column + 1):
                    try:
                        cell_value = data_sheet.cell(row=2, column=col).value
                        if cell_value is not None:
                            float(cell_value)
                            numeric_cols += 1
                    except (ValueError, TypeError):
                        pass
                
                if numeric_cols > 0:
                    insights_sheet['A28'] = f"💰 {numeric_cols} ta raqamli ustun aniqlandi, bu statistik tahlil uchun yaxshi"
                else:
                    insights_sheet['A28'] = "⚠️ Raqamli ma'lumotlar cheklangan, ko'proq raqamli ma'lumotlar qo'shish tavsiya etiladi"
            
            # Add visualization recommendations
            insights_sheet['A30'] = "🎨 5. Vizualizatsiya tavsiyalari:"
            insights_sheet['A30'].font = Font(bold=True, size=14, color="E67E22")
            
            insights_sheet['A32'] = "📊 Bar chartlar kategoriyalarni solishtirish uchun ajoyib"
            insights_sheet['A33'] = "🥧 Pie chartlar nisbatlarni ko'rsatishda foydalidir"
            insights_sheet['A34'] = "📈 Line chartlar vaqt bo'yicha o'zgarishlarni kuzatishda yaxshi"
            insights_sheet['A35'] = "🎨 Conditional formatting ma'lumotlar tendentsiyalarini tezda aniqlashga yordam beradi"
            
            # Style the insights sheet with better formatting
            for row in insights_sheet.iter_rows(min_row=1, max_row=insights_sheet.max_row, min_col=1, max_col=insights_sheet.max_column):
                for cell in row:
                    if cell.row > 1 and cell.value is not None:
                        # Add colorful borders
                        border_colors = ["FFEAA7", "FFD54F", "FFB300", "FFA000"]
                        color_index = (cell.row + cell.column) % len(border_colors)
                        border_color = border_colors[color_index]
                        
                        cell.border = Border(
                            left=Side(style='thin', color=border_color),
                            right=Side(style='thin', color=border_color),
                            top=Side(style='thin', color=border_color),
                            bottom=Side(style='thin', color=border_color)
                        )
                        
                        # Add background color for headers
                        if cell.row in [3, 10, 17, 24, 30]:
                            cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
            
            # Auto-adjust column widths
            for column in insights_sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                insights_sheet.column_dimensions[column_letter].width = min(adjusted_width, 50)
                
        except Exception as e:
            logger.warning(f"Error adding insights: {e}")

    def _apply_professional_styling(self, wb):
        """Apply professional styling to all sheets in the workbook"""
        try:
            # Define professional color schemes
            color_schemes = {
                'data': {
                    'header_bg': '2C3E50',
                    'header_text': 'FFFFFF',
                    'even_row': 'F8F9F9',
                    'odd_row': 'FFFFFF'
                },
                'summary': {
                    'header_bg': '4ECDC4',
                    'header_text': 'FFFFFF'
                },
                'charts': {
                    'header_bg': '96CEB4',
                    'header_text': 'FFFFFF'
                },
                'insights': {
                    'header_bg': 'FFEAA7',
                    'header_text': '2C3E50'
                }
            }
            
            # Apply styling to each sheet
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Apply sheet-specific styling
                if sheet_name == "Ma'lumotlar":
                    self._style_data_sheet(sheet, color_schemes['data'])
                elif sheet_name == "Xulosa":
                    self._style_summary_sheet(sheet, color_schemes['summary'])
                elif sheet_name == "Grafiklar":
                    self._style_charts_sheet(sheet, color_schemes['charts'])
                elif sheet_name == "Tushunchalar":
                    self._style_insights_sheet(sheet, color_schemes['insights'])
                
                # Add sheet tab colors
                if sheet_name == "Ma'lumotlar":
                    sheet.sheet_properties.tabColor = Color("4ECDC4")
                elif sheet_name == "Xulosa":
                    sheet.sheet_properties.tabColor = Color("96CEB4")
                elif sheet_name == "Grafiklar":
                    sheet.sheet_properties.tabColor = Color("45B7D1")
                elif sheet_name == "Tushunchalar":
                    sheet.sheet_properties.tabColor = Color("FFEAA7")
                
        except Exception as e:
            logger.warning(f"Error applying professional styling: {e}")

    def _style_data_sheet(self, sheet, color_scheme):
        """Style the data sheet with professional formatting"""
        # Data sheet is already well-styled in _parse_and_populate_data
        # Add any additional styling here if needed
        pass

    def _style_summary_sheet(self, sheet, color_scheme):
        """Style the summary sheet with professional formatting"""
        # Summary sheet is already well-styled in _add_summary_statistics
        # Add any additional styling here if needed
        pass

    def _style_charts_sheet(self, sheet, color_scheme):
        """Style the charts sheet with professional formatting"""
        # Charts sheet is already well-styled in _add_charts
        # Add any additional styling here if needed
        pass

    def _style_insights_sheet(self, sheet, color_scheme):
        """Style the insights sheet with professional formatting"""
        # Insights sheet is already well-styled in _add_insights
        # Add any additional styling here if needed
        pass